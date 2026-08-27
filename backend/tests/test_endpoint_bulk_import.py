"""Bulk endpoint import (Excel): an admin who already knows a list of
devices (e.g. IT's asset inventory) uploads a filled-in spreadsheet
instead of clicking "Register" 50 times — complementary to enrollment
tokens, which are for devices whose hostname isn't known in advance."""

from __future__ import annotations

import io

import openpyxl
import pytest


def _workbook_bytes(rows: list[list], header: list[str] | None = None) -> bytes:
    header = header or ["name", "hostname", "os", "os_version"]
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_download_import_template_is_a_real_readable_workbook(client, auth_headers):
    response = client.get("/api/v1/endpoints/import-template", headers=auth_headers)
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    header = [cell.value for cell in next(workbook.active.iter_rows(max_row=1))]
    assert header == ["name", "hostname", "os", "os_version"]
    # At least one example row beyond the header, to show the expected shape.
    assert workbook.active.max_row >= 2


def test_import_template_requires_admin(client):
    response = client.get("/api/v1/endpoints/import-template")
    assert response.status_code == 401


def test_bulk_import_creates_every_valid_row(client, auth_headers):
    file_bytes = _workbook_bytes([
        ["Finance-01", "FIN-01", "windows", "11"],
        ["Build-Server", "build-01", "linux", "Ubuntu 24.04"],
    ])
    response = client.post(
        "/api/v1/endpoints/bulk-import", headers=auth_headers,
        files={"file": ("import.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201, response.text
    body = response.json()
    assert body["created"] == 2
    assert body["failed"] == 0
    assert all(row["status"] == "created" for row in body["rows"])
    assert all(row["api_token"].startswith("dsat_") for row in body["rows"])

    endpoints = client.get("/api/v1/endpoints", headers=auth_headers).json()["items"]
    hostnames = {e["hostname"] for e in endpoints}
    assert {"FIN-01", "build-01"} <= hostnames


def test_bulk_import_reports_per_row_errors_without_aborting_the_rest(client, auth_headers):
    file_bytes = _workbook_bytes([
        ["Good One", "good-01", "windows", ""],
        ["Bad OS", "bad-os-01", "solaris", ""],  # invalid os
        ["Missing Hostname", "", "linux", ""],  # missing required field
        ["Good Two", "good-02", "linux", ""],
    ])
    response = client.post(
        "/api/v1/endpoints/bulk-import", headers=auth_headers,
        files={"file": ("import.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201
    body = response.json()
    assert body["created"] == 2
    assert body["failed"] == 2

    by_hostname = {row["hostname"]: row for row in body["rows"]}
    assert by_hostname["good-01"]["status"] == "created"
    assert by_hostname["good-02"]["status"] == "created"
    assert by_hostname["bad-os-01"]["status"] == "error"
    assert "os" in by_hostname["bad-os-01"]["error"].lower()


def test_bulk_import_duplicate_hostname_is_a_row_error(client, auth_headers):
    client.post(
        "/api/v1/endpoints/register", headers=auth_headers,
        json={"name": "Existing", "hostname": "already-here", "os": "linux"},
    )
    file_bytes = _workbook_bytes([["New Name", "already-here", "linux", ""]])
    response = client.post(
        "/api/v1/endpoints/bulk-import", headers=auth_headers,
        files={"file": ("import.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    body = response.json()
    assert body["created"] == 0
    assert body["failed"] == 1
    assert "already registered" in body["rows"][0]["error"]


def test_bulk_import_rejects_a_file_missing_required_columns(client, auth_headers):
    file_bytes = _workbook_bytes([["a", "b"]], header=["foo", "bar"])
    response = client.post(
        "/api/v1/endpoints/bulk-import", headers=auth_headers,
        files={"file": ("import.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 422


def test_bulk_import_rejects_a_non_excel_file_without_crashing(client, auth_headers):
    response = client.post(
        "/api/v1/endpoints/bulk-import", headers=auth_headers,
        files={"file": ("notes.txt", b"this is not an excel file", "text/plain")},
    )
    assert response.status_code == 422


def test_bulk_import_skips_blank_rows(client, auth_headers):
    file_bytes = _workbook_bytes([
        ["Real Row", "real-01", "linux", ""],
        [None, None, None, None],
    ])
    response = client.post(
        "/api/v1/endpoints/bulk-import", headers=auth_headers,
        files={"file": ("import.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    body = response.json()
    assert body["created"] == 1
    assert body["failed"] == 0


def test_bulk_import_requires_admin(client, db_session_factory, org_and_admin):
    import uuid
    from datetime import datetime, timezone

    from datasentinel_backend.models.models import User
    from datasentinel_backend.security.passwords import hash_password

    org, _ = org_and_admin
    session = db_session_factory()
    viewer = User(
        id=uuid.uuid4(), org_id=org.id, email="viewer2@acme-corp.example.com",
        hashed_password=hash_password("hunter2222"), role="viewer", is_active=True,
        created_at=datetime.now(timezone.utc),
    )
    session.add(viewer)
    session.commit()
    session.close()

    login = client.post("/api/v1/auth/login", json={"email": viewer.email, "password": "hunter2222"})
    viewer_headers = {"Authorization": f"Bearer {login.json()['access_token']}"}

    file_bytes = _workbook_bytes([["X", "x-01", "linux", ""]])
    response = client.post(
        "/api/v1/endpoints/bulk-import", headers=viewer_headers,
        files={"file": ("import.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 403


def test_bulk_import_is_audit_logged_with_a_summary_not_raw_tokens(client, auth_headers):
    file_bytes = _workbook_bytes([["Audit Row", "audit-01", "linux", ""]])
    client.post(
        "/api/v1/endpoints/bulk-import", headers=auth_headers,
        files={"file": ("import.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    logs = client.get("/api/v1/audit-logs", headers=auth_headers).json()["items"]
    entry = next(e for e in logs if e["action"] == "endpoints.bulk_imported")
    assert entry["details"]["created"] == 1
    assert "dsat_" not in str(entry)
