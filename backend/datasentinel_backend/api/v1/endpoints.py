"""Endpoint registration and listing. Three ways an endpoint comes into
existence: an authenticated admin registers one directly
(`POST /register`), an agent self-registers using a reusable enrollment
token (`POST /enroll`, spec sections 7-13), or an admin bulk-creates many
at once from a spreadsheet of already-known devices (`POST /bulk-import`).
All three funnel through `services.endpoints.create_endpoint` so they stay
consistent.
"""

from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from datasentinel_backend.api.v1.schemas import (
    BulkImportResponse,
    BulkImportRow,
    EndpointEnrollRequest,
    EndpointRegisterRequest,
    EndpointRegisterResponse,
    EndpointResponse,
    EndpointUpdateRequest,
)
from datasentinel_backend.core.database import get_db
from datasentinel_backend.models.models import EnrollmentToken, Endpoint, Policy, User
from datasentinel_backend.security.dependencies import get_current_user, require_admin
from datasentinel_backend.security.tokens import parse_enrollment_token, verify_api_token
from datasentinel_backend.services.audit import log_action
from datasentinel_backend.services.enrollment import EnrollmentTokenError, validate_enrollment_token
from datasentinel_backend.services.endpoints import DuplicateHostname, compute_risk_score, create_endpoint, get_last_scan_at

router = APIRouter(prefix="/endpoints", tags=["endpoints"])

_IMPORT_TEMPLATE_COLUMNS = ["name", "hostname", "os", "os_version"]


def _to_response(db: Session, endpoint: Endpoint) -> EndpointResponse:
    return EndpointResponse(
        id=endpoint.id,
        name=endpoint.name,
        hostname=endpoint.hostname,
        os=endpoint.os,
        os_version=endpoint.os_version,
        agent_version=endpoint.agent_version,
        status=endpoint.status,
        last_seen_at=endpoint.last_seen_at,
        registered_at=endpoint.registered_at,
        last_scan=get_last_scan_at(db, endpoint.id),
        risk_score=compute_risk_score(db, endpoint.id),
        policy_id=endpoint.policy_id,
    )


@router.post("/register", response_model=EndpointRegisterResponse, status_code=status.HTTP_201_CREATED)
def register_endpoint(
    payload: EndpointRegisterRequest, db: Session = Depends(get_db), user: User = Depends(require_admin)
) -> EndpointRegisterResponse:
    try:
        endpoint, token = create_endpoint(
            db, org_id=user.org_id, name=payload.name, hostname=payload.hostname,
            os=payload.os, os_version=payload.os_version, agent_version=payload.agent_version,
        )
    except DuplicateHostname as exc:
        raise HTTPException(status.HTTP_409_CONFLICT, str(exc)) from exc

    log_action(
        db, org_id=user.org_id, actor_type="user", actor_id=user.id,
        action="endpoint.registered", target_type="endpoint", target_id=endpoint.id,
        details={"hostname": endpoint.hostname, "os": endpoint.os},
    )
    db.commit()
    db.refresh(endpoint)

    return EndpointRegisterResponse(endpoint=_to_response(db, endpoint), api_token=token)


@router.post("/enroll", response_model=EndpointRegisterResponse, status_code=status.HTTP_201_CREATED)
def enroll_endpoint(payload: EndpointEnrollRequest, db: Session = Depends(get_db)) -> EndpointRegisterResponse:
    """Self-registration using a reusable enrollment token — no user JWT
    or endpoint API token involved; the enrollment token itself is the
    only credential. Deliberately unauthenticated at the FastAPI-dependency
    level for that reason, matching how `/auth/login` also takes no prior
    credential."""
    token_id_str = parse_enrollment_token(payload.enrollment_token)
    if token_id_str is None:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Invalid enrollment token")

    try:
        token_id = uuid.UUID(token_id_str)
    except ValueError:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Invalid enrollment token")

    token = db.get(EnrollmentToken, token_id)
    if token is None or not verify_api_token(payload.enrollment_token, token.hashed_token):
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Invalid enrollment token")

    try:
        validate_enrollment_token(token, requested_os=payload.os)
    except EnrollmentTokenError as exc:
        log_action(
            db, org_id=token.org_id, actor_type="enrollment_token", actor_id=token.id,
            action="endpoint.enrollment_rejected", target_type="enrollment_token", target_id=token.id,
            details={"reason": str(exc), "hostname": payload.hostname},
        )
        db.commit()
        raise HTTPException(status.HTTP_403_FORBIDDEN, str(exc)) from exc

    try:
        endpoint, api_token = create_endpoint(
            db, org_id=token.org_id, name=payload.name, hostname=payload.hostname,
            os=payload.os, os_version=payload.os_version, agent_version=payload.agent_version,
            policy_id=token.policy_id,
        )
    except DuplicateHostname as exc:
        raise HTTPException(status.HTTP_409_CONFLICT, str(exc)) from exc

    token.current_uses += 1

    log_action(
        db, org_id=token.org_id, actor_type="enrollment_token", actor_id=token.id,
        action="endpoint.enrolled", target_type="endpoint", target_id=endpoint.id,
        details={"hostname": endpoint.hostname, "os": endpoint.os, "enrollment_token_name": token.name},
    )
    db.commit()
    db.refresh(endpoint)

    return EndpointRegisterResponse(endpoint=_to_response(db, endpoint), api_token=api_token)


@router.get("/import-template")
def download_import_template(user: User = Depends(require_admin)) -> StreamingResponse:
    """An .xlsx an admin fills in and re-uploads to `/bulk-import` — column
    headers must match exactly what that endpoint parses."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Endpoints"
    sheet.append(_IMPORT_TEMPLATE_COLUMNS)
    sheet.append(["Finance-Laptop-01", "FIN-LAPTOP-01", "windows", "11"])
    sheet.append(["Build-Server-01", "build-01", "linux", "Ubuntu 24.04"])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=datasentinel-endpoint-import-template.xlsx"},
    )


@router.post("/bulk-import", response_model=BulkImportResponse, status_code=status.HTTP_201_CREATED)
def bulk_import_endpoints(
    file: UploadFile, db: Session = Depends(get_db), user: User = Depends(require_admin)
) -> BulkImportResponse:
    """Admin-known-device-list onboarding: upload a filled-in copy of the
    `/import-template` spreadsheet, get back one API token per successfully
    created endpoint. Each row is independent — one bad row (missing
    hostname, invalid OS, duplicate hostname) is recorded as an error and
    every other row still gets processed, matching the agent's own
    "one bad file never aborts the scan" philosophy applied to imports.
    """
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file.file.read()), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - any malformed-upload failure must return 422, not 500
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, f"Could not read the uploaded file: {exc}") from exc

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(cell).strip().lower() if cell is not None else "" for cell in next(rows_iter)]
    except StopIteration:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "The uploaded file is empty")

    required = {"name", "hostname", "os"}
    if not required.issubset(header):
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            f"Missing required column(s): {sorted(required - set(header))}. Download the template for the exact format.",
        )
    col = {name: idx for idx, name in enumerate(header)}

    results: list[BulkImportRow] = []
    created_count = 0
    for row_number, raw_row in enumerate(rows_iter, start=2):  # row 1 is the header
        if raw_row is None or all(cell is None for cell in raw_row):
            continue  # skip blank rows rather than erroring on them

        def cell(key: str) -> str | None:
            idx = col.get(key)
            if idx is None or idx >= len(raw_row) or raw_row[idx] is None:
                return None
            return str(raw_row[idx]).strip()

        name, hostname, os_value = cell("name"), cell("hostname"), cell("os")
        os_version = cell("os_version")

        if not name or not hostname or not os_value:
            results.append(BulkImportRow(row=row_number, name=name or "", hostname=hostname or "", status="error", error="name, hostname, and os are all required"))
            continue
        if os_value.lower() not in ("windows", "linux"):
            results.append(BulkImportRow(row=row_number, name=name, hostname=hostname, status="error", error=f"os must be 'windows' or 'linux', got '{os_value}'"))
            continue

        try:
            endpoint, token = create_endpoint(
                db, org_id=user.org_id, name=name, hostname=hostname, os=os_value.lower(), os_version=os_version,
            )
        except DuplicateHostname as exc:
            results.append(BulkImportRow(row=row_number, name=name, hostname=hostname, status="error", error=str(exc)))
            continue

        results.append(BulkImportRow(row=row_number, name=name, hostname=hostname, status="created", api_token=token))
        created_count += 1

    log_action(
        db, org_id=user.org_id, actor_type="user", actor_id=user.id,
        action="endpoints.bulk_imported", target_type=None, target_id=None,
        details={"created": created_count, "failed": len(results) - created_count, "filename": file.filename},
    )
    db.commit()

    return BulkImportResponse(created=created_count, failed=len(results) - created_count, rows=results)


@router.get("", response_model=list[EndpointResponse])
def list_endpoints(db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> list[EndpointResponse]:
    stmt = select(Endpoint).where(Endpoint.org_id == user.org_id).order_by(Endpoint.registered_at.desc())
    endpoints = list(db.execute(stmt).scalars())
    return [_to_response(db, endpoint) for endpoint in endpoints]


@router.get("/{endpoint_id}", response_model=EndpointResponse)
def get_endpoint(
    endpoint_id: uuid.UUID, db: Session = Depends(get_db), user: User = Depends(get_current_user)
) -> EndpointResponse:
    endpoint = db.get(Endpoint, endpoint_id)
    if endpoint is None or endpoint.org_id != user.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Endpoint not found")
    return _to_response(db, endpoint)


@router.patch("/{endpoint_id}", response_model=EndpointResponse)
def update_endpoint(
    endpoint_id: uuid.UUID,
    payload: EndpointUpdateRequest,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> EndpointResponse:
    """Admin-only: assign or clear this endpoint's policy override directly
    (the other way to set it is automatically, via an enrollment token's own
    policy_id at enroll time)."""
    endpoint = db.get(Endpoint, endpoint_id)
    if endpoint is None or endpoint.org_id != user.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Endpoint not found")

    if payload.policy_id is not None:
        policy = db.get(Policy, payload.policy_id)
        if policy is None or policy.org_id != user.org_id:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Policy not found")

    previous_policy_id = endpoint.policy_id
    endpoint.policy_id = payload.policy_id
    db.flush()

    log_action(
        db, org_id=user.org_id, actor_type="user", actor_id=user.id,
        action="endpoint.policy_assigned", target_type="endpoint", target_id=endpoint.id,
        details={"from": str(previous_policy_id) if previous_policy_id else None,
                 "to": str(payload.policy_id) if payload.policy_id else None},
    )
    db.commit()
    db.refresh(endpoint)
    return _to_response(db, endpoint)
